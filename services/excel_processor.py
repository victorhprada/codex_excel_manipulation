"""
Excel Processing Service
Contém toda a lógica de negócio para manipulação de arquivos Excel.
Isolado da camada de apresentação (UI/API).
"""
from __future__ import annotations

from io import BytesIO
from copy import copy
import unicodedata

import pandas as pd
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


# =====================
# Constantes
# =====================
CENTER_SHEET_NAME = "Detalhado"
COLUMN_ESTABELECIMENTO = "ESTABELECIMENTO"
CHECKOUT_COLUMN = "CHECKOUT"

COST_SHEET_NAME = "Custo empresa"
DISCOUNT_SHEET_NAME = "Desconto folha"

COST_FILTER_VALUE = "TARIFA RESGATE LIMITE PARA FLEX"
DISCOUNT_FILTER_VALUE = "RESGATE LIMITE PARA FLEX"

OVERVIEW_SHEET_NAME = "Overview"

# Labels existentes no arquivo ORIGINAL
OVERVIEW_CHECKOUT_PAGAR_LABEL = "Checkouts a pagar"
OVERVIEW_TAXA_ADMIN_LABEL = "Taxa administrativa"
OVERVIEW_SUBSIDIOS_LABEL = "Subsídios"
OVERVIEW_CREDITOS_INSERIDOS_LABEL = "Créditos inseridos"  # Label para remoção

# Labels finais desejados
OVERVIEW_CHECKOUT_FOLHA_LABEL = "Checkouts Folha colab."
OVERVIEW_CHECKOUT_EMPRESA_LABEL = "Checkouts a pagar Empresa"
OVERVIEW_CUSTO_EMPRESA_LABEL = "Custo empresa (Taxa tarifas)"
OVERVIEW_TOTAL_LABEL = "TOTAL DA EMPRESA"
OVERVIEW_A_DEBITAR_LABEL = "A debitar em folha"
OVERVIEW_TOTAL_FUNC_LABEL = "TOTAL DO FUNCIONÁRIO"
OVERVIEW_TOTAL_FECHAMENTO_LABEL = "TOTAL DO FECHAMENTO"

COST_HEADER_ESTABELECIMENTO = "ESTABELECIMENTO"
COST_HEADER_CHECKOUT = "CHECKOUT"
COST_HEADER_DEBITO = "DEBITO EM FOLHA"
COST_HEADER_DEBITO_ACCENT = "DÉBITO EM FOLHA"


# =====================
# Helpers
# =====================
def normalize_text(value: object) -> str:
    """
    Normaliza texto removendo acentos, convertendo para lowercase e removendo espaços.
    """
    if value is None:
        return ""
    text = str(value).strip().lower()
    return "".join(
        c for c in unicodedata.normalize("NFD", text)
        if unicodedata.category(c) != "Mn"
    )


def find_label_cell(sheet, label: str):
    """
    Procura uma célula contendo o label especificado (normalizado).
    """
    target = normalize_text(label)
    for row in sheet.iter_rows():
        for cell in row:
            if normalize_text(cell.value) == target:
                return cell
    return None


def find_value_cell(sheet, label_cell):
    """
    Encontra a célula de valor à direita de uma célula de label.
    """
    if not label_cell:
        return None
    for cell in sheet[label_cell.row]:
        if cell.column > label_cell.column and cell.value not in (None, ""):
            return cell
    return None


def find_header_column(sheet, labels: set[str]) -> str | None:
    """
    Encontra a letra da coluna que contém um dos labels do header.
    """
    normalized = {normalize_text(label) for label in labels}
    for cell in sheet[1]:
        if normalize_text(cell.value) in normalized:
            return cell.column_letter
    return None


def copy_row_style(source_row, target_row) -> None:
    """
    Copia estilos de formatação de uma linha para outra.
    """
    for source_cell, target_cell in zip(source_row, target_row):
        if source_cell.font:
            target_cell.font = copy(source_cell.font)
        if source_cell.fill:
            target_cell.fill = copy(source_cell.fill)
        if source_cell.border:
            target_cell.border = copy(source_cell.border)
        if source_cell.alignment:
            target_cell.alignment = copy(source_cell.alignment)
        target_cell.number_format = source_cell.number_format


def compress_blank_rows_visual(sheet, start_row: int, end_row: int, blank_height: float = 2.0):
    """
    Comprime visualmente linhas em branco reduzindo sua altura.
    """
    for r in range(start_row, end_row + 1):
        row = sheet[r]
        if all(cell.value in (None, "") for cell in row):
            sheet.row_dimensions[r].height = blank_height


# =====================
# Processamento Principal
# =====================
def process_excel(file_bytes: bytes) -> BytesIO:
    """
    Processa um arquivo Excel aplicando regras de negócio específicas.
    
    Args:
        file_bytes: Bytes do arquivo Excel (.xlsx)
        
    Returns:
        BytesIO contendo o arquivo Excel processado
        
    Raises:
        ValueError: Se campos obrigatórios não forem encontrados
        Exception: Para outros erros de processamento
    """
    excel_file = pd.ExcelFile(BytesIO(file_bytes))

    detailed = pd.read_excel(excel_file, sheet_name=CENTER_SHEET_NAME)

    # Máscara: True se tiver checkout (data preenchida), False se vazio
    checkout_filled = (
        detailed[CHECKOUT_COLUMN].notna()
        & detailed[CHECKOUT_COLUMN].astype(str).str.strip().ne("")
    )

    # === BLOCO 1 (TOPO): TARIFA SEM DATA DE CHECKOUT ===
    cost_tarifa_no_checkout = detailed[
        (detailed[COLUMN_ESTABELECIMENTO] == COST_FILTER_VALUE)
        & ~checkout_filled
    ]

    # === BLOCO 2 (MEIO): TARIFA COM DATA DE CHECKOUT ===
    cost_tarifa_checkout = detailed[
        (detailed[COLUMN_ESTABELECIMENTO] == COST_FILTER_VALUE)
        & checkout_filled
    ]

    # === BLOCO 3 (FIM): RESGATE COM DATA DE CHECKOUT ===
    cost_resgate_checkout = detailed[
        (detailed[COLUMN_ESTABELECIMENTO] == DISCOUNT_FILTER_VALUE)
        & checkout_filled
    ]

    # Labels divisores
    title_empresa = pd.DataFrame([{detailed.columns[0]: "Checkouts Empresa"}])
    title_folha = pd.DataFrame([{detailed.columns[0]: "Checkouts Folha colab"}])

    title_empresa = title_empresa.reindex(columns=detailed.columns, fill_value="")
    title_folha = title_folha.reindex(columns=detailed.columns, fill_value="")

    # Montagem final
    cost_frame = pd.concat(
        [
            cost_tarifa_no_checkout, 
            title_empresa, 
            cost_tarifa_checkout, 
            title_folha, 
            cost_resgate_checkout
        ],
        ignore_index=True,
    )

    # Aba Desconto Folha (Resgates sem checkout)
    discount_frame = detailed[
        (detailed[COLUMN_ESTABELECIMENTO] == DISCOUNT_FILTER_VALUE) & ~checkout_filled
    ]

    workbook = load_workbook(BytesIO(file_bytes))
    overview_sheet = workbook[OVERVIEW_SHEET_NAME]

    # === REMOVER LINHA "Créditos inseridos" ===
    creditos_cell = find_label_cell(overview_sheet, OVERVIEW_CREDITOS_INSERIDOS_LABEL)
    if creditos_cell:
        overview_sheet.delete_rows(creditos_cell.row)

    # === Reaproveita linhas base do Overview ===
    checkout_pagar_cell = find_label_cell(overview_sheet, OVERVIEW_CHECKOUT_PAGAR_LABEL)
    taxa_admin_cell = find_label_cell(overview_sheet, OVERVIEW_TAXA_ADMIN_LABEL)
    subsidios_cell = find_label_cell(overview_sheet, OVERVIEW_SUBSIDIOS_LABEL)

    if not checkout_pagar_cell or not taxa_admin_cell or not subsidios_cell:
        raise ValueError("Não foi possível localizar as linhas base do Overview.")

    checkout_pagar_cell.value = OVERVIEW_CHECKOUT_FOLHA_LABEL
    taxa_admin_cell.value = OVERVIEW_CHECKOUT_EMPRESA_LABEL
    subsidios_cell.value = OVERVIEW_CUSTO_EMPRESA_LABEL

    # Mantém o estilo
    copy_row_style(overview_sheet[checkout_pagar_cell.row], overview_sheet[taxa_admin_cell.row])
    copy_row_style(overview_sheet[checkout_pagar_cell.row], overview_sheet[subsidios_cell.row])

    checkout_folha_cell = checkout_pagar_cell
    checkout_empresa_cell = taxa_admin_cell
    custo_empresa_cell = subsidios_cell

    total_empresa_cell = find_label_cell(overview_sheet, OVERVIEW_TOTAL_LABEL)
    a_debitar_cell = find_label_cell(overview_sheet, OVERVIEW_A_DEBITAR_LABEL)
    total_func_cell = find_label_cell(overview_sheet, OVERVIEW_TOTAL_FUNC_LABEL)
    total_fechamento_cell = find_label_cell(overview_sheet, OVERVIEW_TOTAL_FECHAMENTO_LABEL)

    # === Recria abas ===
    for name in (COST_SHEET_NAME, DISCOUNT_SHEET_NAME):
        if name in workbook.sheetnames:
            del workbook[name]

    cost_sheet = workbook.create_sheet(COST_SHEET_NAME)
    for row in dataframe_to_rows(cost_frame, index=False, header=True):
        cost_sheet.append(row)

    discount_sheet = workbook.create_sheet(DISCOUNT_SHEET_NAME)
    for row in dataframe_to_rows(discount_frame, index=False, header=True):
        discount_sheet.append(row)

    # === Fórmulas ===
    cost_debito_col = find_header_column(cost_sheet, {COST_HEADER_DEBITO, COST_HEADER_DEBITO_ACCENT})
    cost_est_col = find_header_column(cost_sheet, {COST_HEADER_ESTABELECIMENTO})
    cost_checkout_col = find_header_column(cost_sheet, {COST_HEADER_CHECKOUT})

    if not (cost_debito_col and cost_est_col and cost_checkout_col):
        raise ValueError("Não foi possível identificar colunas obrigatórias na aba 'Custo empresa'.")

    v_checkout_folha = find_value_cell(overview_sheet, checkout_folha_cell)
    v_checkout_empresa = find_value_cell(overview_sheet, checkout_empresa_cell)
    v_custo_empresa = find_value_cell(overview_sheet, custo_empresa_cell)

    if not (v_checkout_folha and v_checkout_empresa and v_custo_empresa):
        raise ValueError("Não foi possível localizar as células de VALOR no Overview.")

    # 1. Checkout Folha = Soma (RESGATE) Onde (Checkout não é vazio)
    v_checkout_folha.value = (
        f"=SUMIFS('Custo empresa'!{cost_debito_col}:{cost_debito_col},"
        f"'Custo empresa'!{cost_est_col}:{cost_est_col},\"{DISCOUNT_FILTER_VALUE}\","
        f"'Custo empresa'!{cost_checkout_col}:{cost_checkout_col},\"<>\")"
    )

    # 2. Checkout Empresa = Soma (TARIFA) Onde (Checkout não é vazio)
    v_checkout_empresa.value = (
        f"=SUMIFS('Custo empresa'!{cost_debito_col}:{cost_debito_col},"
        f"'Custo empresa'!{cost_est_col}:{cost_est_col},\"{COST_FILTER_VALUE}\","
        f"'Custo empresa'!{cost_checkout_col}:{cost_checkout_col},\"<>\")"
    )

    # 3. Custo empresa = Soma (TARIFA) Onde (Checkout É vazio)
    v_custo_empresa.value = (
        f"=SUMIFS('Custo empresa'!{cost_debito_col}:{cost_debito_col},"
        f"'Custo empresa'!{cost_checkout_col}:{cost_checkout_col},\"=\")"
    )

    # Total empresa
    total_empresa_value = find_value_cell(overview_sheet, total_empresa_cell)
    if not total_empresa_value:
        raise ValueError("Não foi possível localizar a célula de valor de 'TOTAL DA EMPRESA'.")

    total_empresa_value.value = f"=SUM({v_checkout_folha.coordinate},{v_checkout_empresa.coordinate},{v_custo_empresa.coordinate})"

    # A debitar em folha
    a_debitar_value = find_value_cell(overview_sheet, a_debitar_cell)
    if not a_debitar_value:
        raise ValueError("Não foi possível localizar a célula de valor de 'A debitar em folha'.")

    a_debitar_value.value = "=SUM('Desconto folha'!M:M)"

    # Total funcionário
    total_func_value = find_value_cell(overview_sheet, total_func_cell)
    if not total_func_value:
        raise ValueError("Não foi possível localizar a célula de valor de 'TOTAL DO FUNCIONÁRIO'.")

    total_func_value.value = f"={a_debitar_value.coordinate}"

    # Total fechamento
    if not total_fechamento_cell:
        raise ValueError("Não foi possível localizar o label de 'TOTAL DO FECHAMENTO'.")

    total_fechamento_value = overview_sheet.cell(
        row=total_fechamento_cell.row + 1,
        column=total_fechamento_cell.column,
    )
    total_fechamento_value.value = f"={total_empresa_value.coordinate}+{total_func_value.coordinate}"

    # Salvar e Retornar
    output_buffer = BytesIO()
    workbook.save(output_buffer)
    output_buffer.seek(0)  # Garantir que o cursor está no início
    
    return output_buffer
