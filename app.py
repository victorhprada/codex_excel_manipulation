from __future__ import annotations

from io import BytesIO
import unicodedata

import pandas as pd
import streamlit as st
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows


CENTER_SHEET_NAME = "Detalhado"
COLUMN_ESTABELECIMENTO = "ESTABELECIMENTO"
CHECKOUT_COLUMN = "CHECKOUT"

COST_SHEET_NAME = "Custo empresa"
DISCOUNT_SHEET_NAME = "Desconto folha"

COST_FILTER_VALUE = "TARIFA RESGATE LIMITE PARA FLEX"
DISCOUNT_FILTER_VALUE = "RESGATE LIMITE PARA FLEX"

OVERVIEW_SHEET_NAME = "Overview"

# Labels existentes / alvo
OVERVIEW_SECTION_LABEL = "PARTE DA EMPRESA"
OVERVIEW_VALUE_HEADER_LABEL = "VALOR"
OVERVIEW_TOTAL_LABEL = "TOTAL DA EMPRESA"
OVERVIEW_TOTAL_FECHAMENTO_LABEL = "TOTAL DO FECHAMENTO"

# Linhas que já existem no Overview (serão reutilizadas/renomeadas)
OVERVIEW_CHECKOUT_PAGAR_LABEL = "Checkouts a pagar"
OVERVIEW_TAXA_ADMIN_LABEL = "Taxa administrativa"
OVERVIEW_SUBSIDIOS_LABEL = "Subsídios"
OVERVIEW_CREDITOS_LABEL = "Créditos inseridos"  # vamos “limpar” sem deletar linha

# Novos nomes (regras)
OVERVIEW_CHECKOUT_FOLHA_LABEL = "Checkouts Folha colab."
OVERVIEW_CHECKOUT_EMPRESA_LABEL = "Checkouts a pagar Empresa"
OVERVIEW_CUSTO_EMPRESA_LABEL = "Custo empresa (Taxa tarifas)"

OVERVIEW_A_DEBITAR_LABEL = "A debitar em folha"
OVERVIEW_TOTAL_FUNC_LABEL = "TOTAL DO FUNCIONÁRIO"

# Headers na aba Custo empresa
COST_HEADER_ESTABELECIMENTO = "ESTABELECIMENTO"
COST_HEADER_CHECKOUT = "CHECKOUT"
COST_HEADER_DEBITO = "DEBITO EM FOLHA"
COST_HEADER_DEBITO_ACCENT = "DÉBITO EM FOLHA"


def normalize_text(value: object) -> str:
    if value is None:
        return ""
    text = str(value).strip().lower()
    return "".join(
        char for char in unicodedata.normalize("NFD", text) if unicodedata.category(char) != "Mn"
    )


def find_label_cell(sheet, label: str):
    target = normalize_text(label)
    for row in sheet.iter_rows():
        for cell in row:
            if normalize_text(cell.value) == target:
                return cell
    return None


def find_header_column_letter(sheet, labels: set[str]) -> str | None:
    normalized = {normalize_text(label) for label in labels}
    for cell in sheet[1]:
        if normalize_text(cell.value) in normalized:
            return cell.column_letter
    return None


def copy_row_style(source_row, target_row) -> None:
    for source_cell, target_cell in zip(source_row, target_row):
        target_cell.font = source_cell.font
        target_cell.fill = source_cell.fill
        target_cell.border = source_cell.border
        target_cell.alignment = source_cell.alignment
        target_cell.number_format = source_cell.number_format


def get_overview_value_col(overview_sheet):
    """
    Descobre a coluna da tabela de valores do Overview.
    Estratégia: encontrar "PARTE DA EMPRESA" e na MESMA linha achar "VALOR".
    """
    section_cell = find_label_cell(overview_sheet, OVERVIEW_SECTION_LABEL)
    if not section_cell:
        return None

    row = overview_sheet[section_cell.row]
    for cell in row:
        if normalize_text(cell.value) == normalize_text(OVERVIEW_VALUE_HEADER_LABEL):
            return cell.column  # índice numérico
    return None


def get_overview_value_cell(overview_sheet, label_cell, value_col: int):
    """
    Retorna a célula de valor da mesma linha do label, na coluna `value_col`,
    mesmo que esteja vazia.
    """
    if not label_cell or not value_col:
        return None
    return overview_sheet.cell(row=label_cell.row, column=value_col)


def process_excel(uploaded_file: BytesIO) -> BytesIO:
    bytes_data = uploaded_file.getvalue()
    excel_file = pd.ExcelFile(BytesIO(bytes_data))

    if CENTER_SHEET_NAME not in excel_file.sheet_names:
        available = ", ".join(excel_file.sheet_names)
        raise ValueError(f"A aba '{CENTER_SHEET_NAME}' nao foi encontrada. Disponiveis: {available}")

    detailed_frame = pd.read_excel(excel_file, sheet_name=CENTER_SHEET_NAME)

    if COLUMN_ESTABELECIMENTO not in detailed_frame.columns:
        raise ValueError(f"A coluna '{COLUMN_ESTABELECIMENTO}' nao existe na aba '{CENTER_SHEET_NAME}'.")
    if CHECKOUT_COLUMN not in detailed_frame.columns:
        raise ValueError(f"A coluna '{CHECKOUT_COLUMN}' nao existe na aba '{CENTER_SHEET_NAME}'.")

    checkout_filled = (
        detailed_frame[CHECKOUT_COLUMN].notna()
        & detailed_frame[CHECKOUT_COLUMN].astype(str).str.strip().ne("")
    )

    # Base Custo empresa:
    cost_no_checkout = detailed_frame[
        detailed_frame[COLUMN_ESTABELECIMENTO].isin([COST_FILTER_VALUE, DISCOUNT_FILTER_VALUE])
        & ~checkout_filled
    ]
    cost_checkout_empresa = detailed_frame[
        (detailed_frame[COLUMN_ESTABELECIMENTO] == COST_FILTER_VALUE) & checkout_filled
    ]
    cost_checkout_folha = detailed_frame[
        (detailed_frame[COLUMN_ESTABELECIMENTO] == DISCOUNT_FILTER_VALUE) & checkout_filled
    ]

    # Base Desconto folha:
    discount_frame = detailed_frame[
        (detailed_frame[COLUMN_ESTABELECIMENTO] == DISCOUNT_FILTER_VALUE) & ~checkout_filled
    ]

    # “Títulos” dentro da aba Custo empresa (mantemos como antes)
    title_empresa = pd.DataFrame([{detailed_frame.columns[0]: "Checkouts Empresa"}]).reindex(
        columns=detailed_frame.columns, fill_value=""
    )
    title_folha = pd.DataFrame([{detailed_frame.columns[0]: "Checkouts Folha colab"}]).reindex(
        columns=detailed_frame.columns, fill_value=""
    )

    cost_frame = pd.concat(
        [cost_no_checkout, title_empresa, cost_checkout_empresa, title_folha, cost_checkout_folha],
        ignore_index=True,
    )

    workbook = load_workbook(BytesIO(bytes_data))
    overview_sheet = workbook[OVERVIEW_SHEET_NAME] if OVERVIEW_SHEET_NAME in workbook.sheetnames else None

    # Remove e recria as abas tabulares
    for sheet_name in (COST_SHEET_NAME, DISCOUNT_SHEET_NAME):
        if sheet_name in workbook.sheetnames:
            del workbook[sheet_name]

    cost_sheet = workbook.create_sheet(COST_SHEET_NAME)
    for row in dataframe_to_rows(cost_frame, index=False, header=True):
        cost_sheet.append(row)

    discount_sheet = workbook.create_sheet(DISCOUNT_SHEET_NAME)
    for row in dataframe_to_rows(discount_frame, index=False, header=True):
        discount_sheet.append(row)

    # ===== Ajustes do OVERVIEW (sem deletar linhas!) =====
    if overview_sheet:
        value_col = get_overview_value_col(overview_sheet)

        # Encontra as linhas existentes que vamos reutilizar
        label_checkout_pagar = find_label_cell(overview_sheet, OVERVIEW_CHECKOUT_PAGAR_LABEL)
        label_taxa_admin = find_label_cell(overview_sheet, OVERVIEW_TAXA_ADMIN_LABEL)
        label_subsidios = find_label_cell(overview_sheet, OVERVIEW_SUBSIDIOS_LABEL)
        label_creditos = find_label_cell(overview_sheet, OVERVIEW_CREDITOS_LABEL)

        # Template de estilo: usa a linha "Checkouts a pagar" (que já existe e está correta)
        template_row = overview_sheet[label_checkout_pagar.row] if label_checkout_pagar else None

        # Renomeia reaproveitando linhas (sem inserir/remover)
        if label_checkout_pagar:
            label_checkout_pagar.value = OVERVIEW_CHECKOUT_FOLHA_LABEL
        if label_taxa_admin:
            label_taxa_admin.value = OVERVIEW_CHECKOUT_EMPRESA_LABEL
        if label_subsidios:
            label_subsidios.value = OVERVIEW_CUSTO_EMPRESA_LABEL

        # “Remove” Créditos inseridos sem deletar linha: limpa label e valor
        if label_creditos and value_col:
            label_creditos.value = ""
            get_overview_value_cell(overview_sheet, label_creditos, value_col).value = None

        # Copia estilo do template para as linhas reaproveitadas (garante Arial)
        if template_row:
            for label_cell in (label_taxa_admin, label_subsidios):
                if label_cell:
                    copy_row_style(template_row, overview_sheet[label_cell.row])

        # Colunas na aba Custo empresa (detectadas pelo header)
        cost_debito_col = find_header_column_letter(cost_sheet, {COST_HEADER_DEBITO, COST_HEADER_DEBITO_ACCENT})
        cost_estab_col = find_header_column_letter(cost_sheet, {COST_HEADER_ESTABELECIMENTO})
        cost_checkout_col = find_header_column_letter(cost_sheet, {COST_HEADER_CHECKOUT})

        # Células de valor no Overview (mesmo vazias!)
        v_checkout_folha = get_overview_value_cell(overview_sheet, label_checkout_pagar, value_col) if value_col else None
        v_checkout_empresa = get_overview_value_cell(overview_sheet, label_taxa_admin, value_col) if value_col else None
        v_custo_empresa = get_overview_value_cell(overview_sheet, label_subsidios, value_col) if value_col else None

        # Fórmulas dinâmicas (sempre em colunas inteiras, sem ranges fixos)
        if cost_debito_col and cost_estab_col and cost_checkout_col:
            if v_checkout_folha:
                v_checkout_folha.value = (
                    f"=SUMIFS('Custo empresa'!{cost_debito_col}:{cost_debito_col},"
                    f"'Custo empresa'!{cost_estab_col}:{cost_estab_col},"
                    f"\"{DISCOUNT_FILTER_VALUE}\","
                    f"'Custo empresa'!{cost_checkout_col}:{cost_checkout_col},\"<>\")"
                )
            if v_checkout_empresa:
                v_checkout_empresa.value = (
                    f"=SUMIFS('Custo empresa'!{cost_debito_col}:{cost_debito_col},"
                    f"'Custo empresa'!{cost_estab_col}:{cost_estab_col},"
                    f"\"{COST_FILTER_VALUE}\","
                    f"'Custo empresa'!{cost_checkout_col}:{cost_checkout_col},\"<>\")"
                )
            if v_custo_empresa:
                # Somar somente quando CHECKOUT estiver vazio:
                v_custo_empresa.value = (
                    f"=SUMIFS('Custo empresa'!{cost_debito_col}:{cost_debito_col},"
                    f"'Custo empresa'!{cost_checkout_col}:{cost_checkout_col},\"=\")"
                )

        # Total da empresa: soma as 3 células do Overview (coord reais)
        total_empresa_label = find_label_cell(overview_sheet, OVERVIEW_TOTAL_LABEL)
        if total_empresa_label and value_col:
            total_empresa_value = get_overview_value_cell(overview_sheet, total_empresa_label, value_col)
            parts = [c.coordinate for c in (v_checkout_folha, v_checkout_empresa, v_custo_empresa) if c]
            if total_empresa_value and parts:
                total_empresa_value.value = f"=SUM({','.join(parts)})"

        # A debitar em folha e total do funcionário
        a_debitar_label = find_label_cell(overview_sheet, OVERVIEW_A_DEBITAR_LABEL)
        total_func_label = find_label_cell(overview_sheet, OVERVIEW_TOTAL_FUNC_LABEL)
        a_debitar_value = get_overview_value_cell(overview_sheet, a_debitar_label, value_col) if (a_debitar_label and value_col) else None
        total_func_value = get_overview_value_cell(overview_sheet, total_func_label, value_col) if (total_func_label and value_col) else None

        if a_debitar_value:
            a_debitar_value.value = "=SUM('Desconto folha'!M:M)"
        if total_func_value and a_debitar_value:
            total_func_value.value = f"={a_debitar_value.coordinate}"

        # Total do fechamento: célula abaixo do label (B8:C8 mesclada normalmente)
        total_fechamento_label = find_label_cell(overview_sheet, OVERVIEW_TOTAL_FECHAMENTO_LABEL)
        if total_fechamento_label:
            # valor fica na linha abaixo, mesma coluna do label
            total_fechamento_value = overview_sheet.cell(
                row=total_fechamento_label.row + 1,
                column=total_fechamento_label.column,
            )
            # pega total empresa e total funcionário
            total_empresa_label = find_label_cell(overview_sheet, OVERVIEW_TOTAL_LABEL)
            total_empresa_value = get_overview_value_cell(overview_sheet, total_empresa_label, value_col) if (total_empresa_label and value_col) else None
            if total_fechamento_value and total_empresa_value and total_func_value:
                total_fechamento_value.value = f"={total_empresa_value.coordinate}+{total_func_value.coordinate}"

    output_buffer = BytesIO()
    workbook.save(output_buffer)
    output_buffer.seek(0)
    return output_buffer


def main() -> None:
    st.title("Gerador de Relatorio Excel")
    st.write(
        "Envie um arquivo Excel (.xlsx) com as abas originais e gere um novo arquivo "
        "com as abas adicionais 'Custo empresa' e 'Desconto folha'."
    )

    uploaded_file = st.file_uploader(
        "Selecione o arquivo Excel (.xlsx)",
        type=["xlsx"],
    )

    if uploaded_file is None:
        st.info("Nenhum arquivo carregado. Envie um arquivo Excel para continuar.")
        return

    if st.button("Processar arquivo"):
        try:
            output_buffer = process_excel(uploaded_file)
        except ValueError as exc:
            st.error(str(exc))
            return
        except Exception as exc:  # noqa: BLE001
            st.error(f"Erro inesperado ao processar o arquivo: {exc}")
            return

        output_filename = f"processado_{uploaded_file.name}"
        st.success("Arquivo processado com sucesso.")
        st.download_button(
            label="Baixar arquivo processado",
            data=output_buffer,
            file_name=output_filename,
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


if __name__ == "__main__":
    main()
